import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# ======================
# CONFIGURATION
# ======================
mstr_json = "MSTR.json"
powerbi_json = "PBI.json"
output_dir = "comparisons"   # root output directory
highlight_color = "FFFF0000"  # Red
MATCHING_THRESHOLD = 50  # %

os.makedirs(output_dir, exist_ok=True)

# ======================
# FUNCTIONS
# ======================
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Convert to string, strip spaces, lowercase everything."""
    return df.astype(str).apply(lambda col: col.str.strip().str.lower())

def compute_row_hash(df: pd.DataFrame) -> pd.Series:
    """Generate unique hash per row."""
    return pd.util.hash_pandas_object(df, index=False).astype(str)

def get_match_stats(row: pd.Series, df_cmp: pd.DataFrame, cols: list):
    """Compare one row against all rows of another DataFrame."""
    row_vals = row[cols].to_numpy()
    cmp_vals = df_cmp[cols].to_numpy()
    matches = (cmp_vals == row_vals)

    match_counts = matches.sum(axis=1)
    if match_counts.max() == 0:
        return None, [], 0, 0

    best_idx = match_counts.argmax()
    best_score = match_counts[best_idx]
    total_cols = len(cols)
    match_pct = (best_score / total_cols) * 100
    matching_cols = np.where(matches[best_idx])[0].tolist()

    return best_idx, matching_cols, best_score, match_pct

def categorize_rows(df1, df2, cols):
    """Classify rows as Matched / Partial / Not Matched with Debug IDs."""
    df1_hashes, df2_hashes = set(df1['row_hash']), set(df2['row_hash'])
    df1_status, df2_status, df1_debug, df2_debug = [], [], [], []
    partial_counter = 1  # Counter for unique Partial Match IDs
    partial_map = {}     # To map df2 indices to debug IDs

    # Process df1
    for i, row in df1.iterrows():
        if row['row_hash'] in df2_hashes:
            df1_status.append("Matched")
            df1_debug.append(None)
        else:
            best_idx, _, score, pct = get_match_stats(row, df2, cols)
            if score == 0 or (100 - pct) >= MATCHING_THRESHOLD:
                df1_status.append("Not Matched")
                df1_debug.append(None)
            else:
                df1_status.append("Partial Matched")
                debug_id = f"PM_{partial_counter:04d}"
                df1_debug.append(debug_id)
                partial_map[best_idx] = debug_id
                partial_counter += 1

    # Process df2
    for j, row in df2.iterrows():
        if row['row_hash'] in df1_hashes:
            df2_status.append("Matched")
            df2_debug.append(None)
        else:
            best_idx, _, score, pct = get_match_stats(row, df1, cols)
            if score == 0 or (100 - pct) >= MATCHING_THRESHOLD:
                df2_status.append("Not Matched")
                df2_debug.append(None)
            else:
                df2_status.append("Partial Matched")
                debug_id = partial_map.get(j, f"PM_{partial_counter:04d}")
                df2_debug.append(debug_id)
                if j not in partial_map:  # only increment if new mapping
                    partial_counter += 1

    return df1_status, df2_status, df1_debug, df2_debug

def highlight_cells(ws, df_orig, df_clean, df_cmp, cols):
    """Apply red highlighting to mismatched cells or rows."""
    highlight = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    for i, row in enumerate(df_orig.itertuples(index=False), start=2):
        if row.Status in ['Not Matched', 'Partial Matched']:
            best_idx, matching_cols, _, _ = get_match_stats(df_clean.iloc[i-2], df_cmp, cols)

            if row.Status == "Partial Matched" and best_idx is not None:
                for col_idx in range(len(cols)):
                    if col_idx not in matching_cols:
                        ws.cell(row=i, column=col_idx+1).fill = highlight
            else:
                for j in range(1, len(cols)+1):
                    ws.cell(row=i, column=j).fill = highlight

def compare_files(file1, file2, output_file):
    """Compare two Excel files and generate a comparison report."""
    print(f"Comparing {file1} vs {file2} -> {output_file}")

    # Read both Excel files
    df1, df2 = pd.read_excel(file1), pd.read_excel(file2)

    # ✅ Sort both DataFrames by all columns
    df1 = df1.sort_values(by=df1.columns.tolist()).reset_index(drop=True)
    df2 = df2.sort_values(by=df2.columns.tolist()).reset_index(drop=True)

    # Clean and compute row hashes
    df1_clean, df2_clean = clean_dataframe(df1.copy()), clean_dataframe(df2.copy())
    df1_clean['row_hash'] = compute_row_hash(df1_clean)
    df2_clean['row_hash'] = compute_row_hash(df2_clean)

    # ✅ Add row_hash into original DataFrames
    df1['row_hash'] = df1_clean['row_hash']
    df2['row_hash'] = df2_clean['row_hash']

    # Columns excluding the row_hash
    cols = [c for c in df1_clean.columns if c != 'row_hash']

    # Add Status + Debug IDs
    df1['Status'], df2['Status'], df1['Debug_ID'], df2['Debug_ID'] = categorize_rows(df1_clean, df2_clean, cols)

    # ✅ Reorder columns to place row_hash, Debug_ID, then Status
    col_order_1 = [c for c in df1.columns if c not in ['Status', 'Debug_ID']] + ['Debug_ID', 'Status']
    col_order_2 = [c for c in df2.columns if c not in ['Status', 'Debug_ID']] + ['Debug_ID', 'Status']
    df1, df2 = df1[col_order_1], df2[col_order_2]

    # Write results into Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="MSTR_File", index=False)
        df2.to_excel(writer, sheet_name="PBI_File", index=False)

    # Highlight mismatches
    wb = load_workbook(output_file)
    highlight_cells(wb["MSTR_File"], df1, df1_clean, df2_clean, cols)
    highlight_cells(wb["PBI_File"], df2, df2_clean, df1_clean, cols)
    wb.save(output_file)

    return output_file

# ======================
# MAIN
# ======================
print("Loading mappings...")
with open(mstr_json) as f:
    mstr_map = json.load(f)["MSTR"]
with open(powerbi_json) as f:
    powerbi_map = json.load(f)["PowerBI"]

# Root dictionary
comparison_results = {"Comparison": {}}

print("Starting comparisons...")
for report, pages in mstr_map.items():
    comparison_results["Comparison"][report] = {}
    for page, visuals in pages.items():
        comparison_results["Comparison"][report][page] = {}
        for visual, mstr_file in visuals.items():
            if report in powerbi_map and page in powerbi_map[report] and visual in powerbi_map[report][page]:
                pbi_file = powerbi_map[report][page][visual]

                # ✅ Create nested directories for Report/Page
                output_subdir = os.path.join(output_dir, report, page)
                os.makedirs(output_subdir, exist_ok=True)

                # ✅ Output file path
                output_path = os.path.join(output_subdir, f"{visual}_comparison.xlsx")

                compare_files(mstr_file, pbi_file, output_path)

                # ✅ Store in JSON structure
                comparison_results["Comparison"][report][page][visual] = output_path

# Save mapping of results
with open("comparison_results.json", "w") as f:
    json.dump(comparison_results, f, indent=2)

print("\n✅ All comparisons done. Results mapping saved in comparison_results.json")