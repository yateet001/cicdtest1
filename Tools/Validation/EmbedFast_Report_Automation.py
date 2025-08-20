import sys
import os
import asyncio
import logging
import re
import json
from typing import Dict, List, Set, Tuple, Optional
from pathlib import Path
from playwright.async_api import async_playwright, Page, Download, Frame, TimeoutError as PWTimeout
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
# ---- Comparison settings / roots ----
OUTPUT_ROOT = os.environ.get("POWERBI_OUTPUT_DIR", "PowerBI_reports")   # already used for PBI output root
MSTR_OUTPUT_ROOT = os.environ.get("MSTR_OUTPUT_DIR", "MSTR_reports")    # parallel root for MSTR
output_dir = "comparisons"
highlight_color = "FFFF0000"  # Red
MATCHING_THRESHOLD = 50       # %
os.makedirs(output_dir, exist_ok=True)

def _resolve_edge_profile_dir():
    """
    Returns the path for Edge profile directory
    """
    PROFILE_DIR = ".edge-user-data"  # persistent profile (reuses your SSO)
    profile_path = str(Path(PROFILE_DIR).resolve())
    Path(profile_path).mkdir(parents=True, exist_ok=True)
    return profile_path

def likely_auth_url(u: str) -> bool:
    """Check if URL is likely an authentication endpoint"""
    u = (u or "").lower()
    return any(s in u for s in [
        "login.microsoftonline.com", "login.microsoft.com",
        "sts.", "adfs.", "sso.", "auth."
    ])

# -------------------- UTIL -------------------- #
def _sanitize_filename(s: str) -> str:
    return (
        (s or "")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
        .replace("?", "_")
        .replace("*", "_")
        .replace("|", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace(" ", "_")
        .strip("_")
    )

def _norm(s: str) -> str:
    """Normalize for matching (case-insensitive, collapse whitespace)."""
    return re.sub(r"\s+", " ", (s or "").strip()).lower()

# ======================
# COMPARISON UTIL FUNCTIONS (as-provided)
# ======================
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize all columns: strip spaces, lowercase, convert to str."""
    return df.astype(str).apply(lambda col: col.str.strip().str.lower())

def compute_row_hash(df: pd.DataFrame) -> pd.Series:
    """Generate deterministic hash per row."""
    return pd.util.hash_pandas_object(df, index=False).astype(str)

def get_best_match(row: pd.Series, df_cmp: pd.DataFrame, cols: list):
    """Find best matching row in df_cmp for given row using column-wise comparison."""
    row_vals = row[cols].to_numpy(copy=False)
    cmp_vals = df_cmp[cols].to_numpy(copy=False)
    matches = (cmp_vals == row_vals)

    match_counts = matches.sum(axis=1)
    if not match_counts.any():
        return None, [], 0, 0

    best_idx = int(match_counts.argmax())
    best_score = match_counts[best_idx]
    total_cols = len(cols)
    match_pct = (best_score / total_cols) * 100
    matching_cols = np.flatnonzero(matches[best_idx]).tolist()

    return best_idx, matching_cols, best_score, match_pct

def categorize_rows(df1, df2, cols):
    """Classify rows into Matched / Partial / Not Matched with Debug IDs."""
    df1_hashes, df2_hashes = set(df1['row_hash']), set(df2['row_hash'])

    df1_status, df1_debug = [], []
    df2_status, df2_debug = [], []

    partial_counter = 1
    partial_map = {}

    # ---- Process df1 ----
    for i, row in df1.iterrows():
        if row['row_hash'] in df2_hashes:
            df1_status.append("Matched")
            df1_debug.append(None)
            continue

        best_idx, _, score, pct = get_best_match(row, df2, cols)
        if score == 0 or (100 - pct) > MATCHING_THRESHOLD:
            df1_status.append("Not Matched")
            df1_debug.append(None)
        else:
            debug_id = f"PM_{partial_counter:04d}"
            df1_status.append("Partial Matched")
            df1_debug.append(debug_id)
            partial_map[best_idx] = debug_id
            partial_counter += 1

    # ---- Process df2 ----
    for j, row in df2.iterrows():
        if row['row_hash'] in df1_hashes:
            df2_status.append("Matched")
            df2_debug.append(None)
            continue

        best_idx, _, score, pct = get_best_match(row, df1, cols)
        if score == 0 or (100 - pct) > MATCHING_THRESHOLD:
            df2_status.append("Not Matched")
            df2_debug.append(None)
        else:
            debug_id = partial_map.get(j, f"PM_{partial_counter:04d}")
            df2_status.append("Partial Matched")
            df2_debug.append(debug_id)
            if j not in partial_map:
                partial_counter += 1

    return df1_status, df2_status, df1_debug, df2_debug

def highlight_cells(ws, df_orig, df_clean, df_cmp, cols):
    """Highlight mismatched cells in Excel sheet."""
    highlight = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    for i, row in enumerate(df_orig.itertuples(index=False), start=2):
        if row.Status not in ["Not Matched", "Partial Matched"]:
            continue

        best_idx, matching_cols, _, _ = get_best_match(df_clean.iloc[i - 2], df_cmp, cols)
        if row.Status == "Partial Matched" and best_idx is not None:
            for col_idx in range(len(cols)):
                if col_idx not in matching_cols:
                    ws.cell(row=i, column=col_idx + 1).fill = highlight
        else:
            for j in range(1, len(cols) + 1):
                ws.cell(row=i, column=j).fill = highlight

def compare_files(file1, file2, output_file):
    """Compare two Excel files and save annotated comparison result."""
    print(f"Comparing {file1} vs {file2} -> {output_file}")

    # Read files
    df1, df2 = pd.read_excel(file1), pd.read_excel(file2)

    # Sort for consistency
    sort_cols = df1.columns.tolist()
    df1, df2 = df1.sort_values(by=sort_cols).reset_index(drop=True), df2.sort_values(by=sort_cols).reset_index(drop=True)

    # Clean + Hash
    df1_clean, df2_clean = clean_dataframe(df1.copy()), clean_dataframe(df2.copy())
    df1_clean['row_hash'], df2_clean['row_hash'] = compute_row_hash(df1_clean), compute_row_hash(df2_clean)
    df1['row_hash'], df2['row_hash'] = df1_clean['row_hash'], df2_clean['row_hash']

    # Columns excluding row_hash
    cols = [c for c in df1_clean.columns if c != 'row_hash']

    # Status + Debug IDs
    df1['Status'], df2['Status'], df1['Debug_ID'], df2['Debug_ID'] = categorize_rows(df1_clean, df2_clean, cols)

    # Reorder columns (row_hash + Debug_ID + Status at end)
    def reorder(df):
        base_cols = [c for c in df.columns if c not in ['Debug_ID', 'Status']]
        return df[base_cols + ['Debug_ID', 'Status']]

    df1, df2 = reorder(df1), reorder(df2)

    # Write results
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="MSTR_File", index=False)
        df2.to_excel(writer, sheet_name="PBI_File", index=False)

    # Highlight mismatches
    wb = load_workbook(output_file)
    highlight_cells(wb["MSTR_File"], df1, df1_clean, df2_clean, cols)
    highlight_cells(wb["PBI_File"], df2, df2_clean, df1_clean, cols)
    wb.save(output_file)

    return output_file

# -------------------- RUNNER (MULTI-REPORT) -------------------- #
class ReportsRunner:
    def __init__(self):
        cfg_path = os.environ.get("REPORT_CONFIG_FILE", "reports.json")
        if not os.path.isfile(cfg_path):
            raise FileNotFoundError(
                f"Config file '{cfg_path}' not found. Set REPORT_CONFIG_FILE or create reports.json."
            )
        with open(cfg_path, "r", encoding="utf-8") as f:
            raw = json.load(f)

        # Support both new and old keys
        all_reports = raw.get("PowerBI_reports") or raw.get("reports") or []
        if not all_reports:
            raise ValueError("Config has no 'PowerBI_reports' or 'reports' entries.")

        desired = os.environ.get("REPORT_NAME")
        if desired:
            self.reports = [r for r in all_reports if _norm(r.get("name")) == _norm(desired)]
            if not self.reports:
                raise ValueError(f"REPORT_NAME='{desired}' not found in config.")
        else:
            self.reports = all_reports  # iterate all

    async def run(self):
        async with async_playwright() as p:
            profile_path = _resolve_edge_profile_dir()
            logger.info(f"Using persistent profile at: {profile_path}")

            try:
                # Launch Edge with persistent profile but allowing new tabs
                context = await p.chromium.launch_persistent_context(
                    user_data_dir=profile_path,
                    channel="msedge",
                    headless=False,
                    viewport={"width": 1920, "height": 1080},
                    accept_downloads=True,
                    ignore_https_errors=True,
                    args=[
                        "--no-first-run",
                        "--no-default-browser-check",
                    ],
                )
                logger.info(f"Launched Edge with profile at '{profile_path}'")
            except Exception as e:
                raise RuntimeError(
                    "Could not launch Microsoft Edge with persistent profile.\n"
                    f"Underlying error: {e}"
                )

            page = context.pages[0] if context.pages else await context.new_page()

            for idx, report in enumerate(self.reports):
                name = report.get("name") or f"report_{idx+1}"
                url = report.get("url")
                pages = report.get("pages", [])

                if not url:
                    logger.warning(f"Report '{name}' missing URL. Skipping.")
                    continue

                logger.info(f"=== [{idx+1}/{len(self.reports)}] Opening report: {name} ===")
                await page.goto(url, wait_until="domcontentloaded")

                # Handle SSO if needed
                try:
                    await page.wait_for_load_state("domcontentloaded", timeout=20000)
                except PWTimeout:
                    pass

                if likely_auth_url(page.url):
                    logger.info("Sign-in detected. Waiting for auth completion...")
                    try:
                        not_auth_pattern = re.compile(
                            r"^(?!.*(login\.microsoftonline\.com|login\.microsoft\.com|sts\.|adfs\.|sso\.|auth\.)).*",
                            re.I
                        )
                        await page.wait_for_url(not_auth_pattern, timeout=240000)
                        logger.info("Authentication completed.")
                    except PWTimeout:
                        logger.warning("Authentication timeout. Please complete login manually.")
                        continue

                # Wait for report to load and handle iframes
                try:
                    
                    try:
                        prompt_msg = (
                            "\nPress ENTER when the report is loaded completely to proceed… "
                        )
                        # Use a non-blocking approach for async contexts
                        loop = asyncio.get_running_loop()
                        await loop.run_in_executor(None, lambda: input(prompt_msg))
                    except EOFError:
                        # If no stdin (e.g., non-interactive run), don’t fail—just add a small grace delay
                        logger.info("stdin not available; continuing automatically in 5 seconds.")
                        await page.wait_for_timeout(5000)
                    # Build plan for pages and visuals
                    def _to_visuals_set(vlist):
                        # Treat "ALL" (any case, extra spaces ok) as wildcard
                        return {"__ALL__"} if any(_norm(v) == "all" for v in (vlist or [])) else {_norm(v) for v in (vlist or [])}

                    pages_order = [p.get("name", f"page_{i+1}") for i, p in enumerate(pages)]
                    page_visuals = {
                        _norm(p.get("name", f"page_{i+1}")): _to_visuals_set(p.get("visuals", []))
                        for i, p in enumerate(pages)
                    }


                    await page.wait_for_selector(
                        "iframe, #pvExplorationHost, [data-testid='artifact-info-title']",
                    )

                    # Check if report is in iframe and switch to it if needed
                    iframe = await page.query_selector("iframe")
                    if iframe:
                        logger.info("Report detected in iframe, switching context...")
                        frame = await iframe.content_frame()
                        if frame:
                            # Store both page and frame - we need both
                            worker = SingleReportWorker(
                                page=page,
                                frame=frame,  # Pass the frame separately
                                config_report_name=name,
                                pages_order=pages_order,
                                page_visuals=page_visuals
                            )
                            await worker.run_for_current_report(url)
                            continue  # Skip the normal flow since we handled it in iframe
                        else:
                            logger.warning("Failed to switch to iframe context")
                            continue
                    
                except PWTimeout:
                    logger.warning("Report surface not detected. Please check if report loaded correctly.")
                    continue

                # Build plan for pages and visuals
                pages_order = [p.get("name", f"page_{i+1}") for i, p in enumerate(pages)]
                page_visuals = {
                    _norm(p.get("name", f"page_{i+1}")): {_norm(v) for v in p.get("visuals", [])}
                    for i, p in enumerate(pages)
                }

                worker = SingleReportWorker(
                    page=page,
                    config_report_name=name,
                    pages_order=pages_order,
                    page_visuals=page_visuals
                )
                await worker.run_for_current_report(url)

            await context.close()
            # ---- Build mapping JSONs then run comparisons (uses the comparison code logic as-is) ----
            # ---- Directory-based comparison (no JSON maps) ----
            pbi_root = os.path.join(os.path.abspath(os.getcwd()), OUTPUT_ROOT)       # e.g., "PowerBI_reports"
            mstr_root = os.path.join(os.path.abspath(os.getcwd()), MSTR_OUTPUT_ROOT) # e.g., "MSTR_reports"

            def _scan_root(root_dir: str):
                """
                Build mapping: report -> page -> visual_key(lowercase file base) -> absolute file path
                Expects: <root>/<Report>/<Page>/data/<VisualBase>.xlsx
                """
                mapping = {}
                if not os.path.isdir(root_dir):
                    return mapping
                for report in os.listdir(root_dir):
                    rep_path = os.path.join(root_dir, report)
                    if not os.path.isdir(rep_path):
                        continue
                    mapping.setdefault(report, {})
                    for page in os.listdir(rep_path):
                        page_path = os.path.join(rep_path, page)
                        data_dir = os.path.join(page_path, "data")
                        if not os.path.isdir(data_dir):
                            continue
                        mapping[report].setdefault(page, {})
                        for fn in os.listdir(data_dir):
                            if fn.lower().endswith((".xlsx", ".csv")):
                                base = os.path.splitext(fn)[0]
                                key = base.lower()
                                mapping[report][page][key] = os.path.join(data_dir, fn)
                return mapping

            # Scan both trees
            pbi_map  = _scan_root(pbi_root)
            mstr_map = _scan_root(mstr_root)

            comparison_results = {"Comparison": {}}
            print("Starting comparisons (directory-based)...")

            for report, pages in mstr_map.items():
                if report not in pbi_map:
                    continue
                comparison_results["Comparison"][report] = {}
                for page, visuals in pages.items():
                    if page not in pbi_map[report]:
                        continue
                    comparison_results["Comparison"][report][page] = {}

                    # Match by lowercase file base (ensures case-insensitive match)
                    for visual_key, mstr_file in visuals.items():
                        if visual_key in pbi_map[report][page]:
                            pbi_file = pbi_map[report][page][visual_key]

                            # Output path: comparisons/<Report>/<Page>/<Visual>_comparison.xlsx
                            output_subdir = os.path.join(output_dir, report, page)
                            os.makedirs(output_subdir, exist_ok=True)

                            # Use the original MSTR visual filename base for readability
                            visual_name = os.path.splitext(os.path.basename(mstr_file))[0]
                            output_path = os.path.join(output_subdir, f"{visual_name}_comparison.xlsx")

                            compare_files(mstr_file, pbi_file, output_path)
                            comparison_results["Comparison"][report][page][visual_name] = output_path

            # Save JSON result mapping
            with open("comparison_results.json", "w", encoding="utf-8") as f:
                json.dump(comparison_results, f, indent=2)

            print("\n✅ All comparisons done (directory-based). Results mapping saved in comparison_results.json")




# -------------------- SINGLE REPORT WORKER -------------------- #
class SingleReportWorker:
    def __init__(self, page: Page, config_report_name: str, pages_order: List[str], 
                 page_visuals: Dict[str, Set[str]], frame: Frame = None):
        self.page = page
        self.frame = frame  # Store the frame if we're working with an iframe
        self.context = frame if frame else page  # Use frame for element operations if available
        self.config_report_name = config_report_name
        self.pages_order = pages_order or []
        self.page_visuals = page_visuals or {}
        self.report_name: Optional[str] = None
        self.download_dir: Optional[str] = None

    async def run_for_current_report(self, url: str):
        await self._setup_report_folder()  # uses config name
        # Try to get pages nav
        try:
            await self.context.wait_for_selector('[data-testid="pages-navigation-list"]', timeout=30000)
            pages_pane = await self.context.query_selector('[data-testid="pages-navigation-list"]')
            page_items = await pages_pane.query_selector_all('[data-testid="pages-navigation-list-items"]')
        except Exception as e:
            logger.warning(f"Could not find pages navigation list: {e}. Will operate on current page.")
            page_items = []

        if self.pages_order and page_items:
            # Map normalized aria-label -> element
            tabs: List[Tuple[object, str]] = []
            for it in page_items:
                label = await it.get_attribute("aria-label") or ""
                label_clean = re.sub(r'\s*selected\s*$', '', label, flags=re.IGNORECASE).strip()
                tabs.append((it, label_clean))

            for cfg_page_name in self.pages_order:
                norm_cfg = _norm(cfg_page_name)
                safe_page_name = _sanitize_filename(cfg_page_name)

                # find matching tab
                match_el = None
                for it, lab in tabs:
                    if _norm(lab) == norm_cfg:
                        match_el = it
                        break
                if not match_el:
                    logger.warning(f"Page '{cfg_page_name}' from config not found in the report; skipping.")
                    continue

                logger.info(f"[{self.config_report_name}] Switching to page: {cfg_page_name}")
                try:
                    await match_el.click()
                    await self.context.wait_for_timeout(1500)
                except Exception as e:
                    logger.warning(f"Failed to click page '{cfg_page_name}': {e}")
                    continue

                await self._capture_page_screenshot(safe_page_name)
                allowed_visuals = self.page_visuals.get(norm_cfg, set())
                await self._export_visuals_on_current_page(safe_page_name, allowed_visuals)

        else:
            # Single-page or tabs not available—use union of visuals across config
            logger.info(f"[{self.config_report_name}] Operating on current page (no tabs or no pages in config).")
            await self._capture_page_screenshot("current_page")
            allowed = set()
            for vset in self.page_visuals.values():
                allowed |= vset
            await self._export_visuals_on_current_page("current_page", allowed)

    async def _setup_report_folder(self):
        safe_report_name = _sanitize_filename(self.config_report_name) or "PowerBI_Report"
        self.report_name = safe_report_name

        root = os.path.join(os.path.abspath(os.getcwd()), OUTPUT_ROOT)
        os.makedirs(root, exist_ok=True)

        self.download_dir = os.path.join(root, self.report_name)
        os.makedirs(self.download_dir, exist_ok=True)
        logger.info(f"Output root: {root} | Report folder: {self.download_dir}")

    async def _capture_page_screenshot(self, safe_page_name: str):
        screenshot_dir = os.path.join(self.download_dir, safe_page_name, "screenshot")
        os.makedirs(screenshot_dir, exist_ok=True)
        screenshot_path = os.path.join(screenshot_dir, f"{safe_page_name}.png")
        try:
            # Always use self.page for screenshots, even when working with frames
            await self.page.screenshot(path=screenshot_path, full_page=True)
            logger.info(f"Screenshot saved: {screenshot_path}")
        except Exception as e:
            logger.error(f"Failed to save screenshot for {safe_page_name}: {e}")

    async def _get_visual_title_for_matching(self, container) -> str:
        # Try header title attribute
        try:
            title_el = await container.query_selector("div[data-testid='visual-title']")
            if title_el:
                t = await title_el.get_attribute("title")
                if t and t.strip():
                    return t.strip()
        except:
            pass

        # Fallback aria-label
        try:
            al = await container.get_attribute("aria-label")
            if al and al.strip():
                return al.strip()
        except:
            pass

        return ""

    async def _get_visual_name_for_files(self, container, index: int) -> str:
        name = ""
        try:
            title = await container.query_selector("div[data-testid='visual-title']")
            if title:
                t = await title.get_attribute("title")
                if t:
                    name = t.strip()
        except:
            pass

        if not name:
            name = (await container.get_attribute("aria-label") or "").strip()

        if not name:
            rd = (await container.get_attribute("aria-roledescription") or "visual").strip()
            name = f"{rd}_{index+1}"

        return _sanitize_filename(name)

    async def _open_menu_for_container(self, container, retries=3):
        for _ in range(retries):
            try:
                await container.scroll_into_view_if_needed()
                await container.hover()
                await self.context.wait_for_timeout(200)

                # Try different button selectors
                button_selectors = [
                    "button[aria-label*='More options']",
                    "button[data-testid='visual-more-options-btn']",
                    ".vcMenuBtn",
                    "[aria-label*='More options']",
                    "[title*='More options']",
                    "[class*='menu-btn']",
                    "[class*='more-options']"
                ]

                more_btn = None
                for selector in button_selectors:
                    more_btn = await container.query_selector(selector)
                    if more_btn and await more_btn.is_visible():
                        break

                if not more_btn:
                    await self.context.wait_for_timeout(300)
                    # Try moving viewport slightly
                    try:
                        await self.page.mouse.wheel(0, 100)
                    except:
                        # If mouse wheel fails, try scrolling the container
                        await container.evaluate("el => el.scrollIntoView({behavior: 'smooth', block: 'center'})")
                    await self.context.wait_for_timeout(200)
                    continue

                # Try different click methods
                try:
                    await more_btn.click()
                except:
                    try:
                        await self.context.evaluate("(b) => b.click()", more_btn)
                    except:
                        try:
                            # Force click using JavaScript
                            await self.context.evaluate("""(element) => {
                                const clickEvent = new MouseEvent('click', {
                                    bubbles: true,
                                    cancelable: true,
                                    view: window
                                });
                                element.dispatchEvent(clickEvent);
                            }""", more_btn)
                        except:
                            continue

                # Check for menu using multiple selectors
                menu_selectors = [
                    "[role='menu']", 
                    ".pbi-menu",
                    "[class*='menu-container']",
                    "[class*='context-menu']"
                ]

                menu = None
                for selector in menu_selectors:
                    try:
                        await self.context.wait_for_selector(selector, timeout=3000)
                        menus = await self.context.query_selector_all(selector)
                        for m in menus[::-1]:
                            if await m.is_visible():
                                menu = m
                                break
                        if menu:
                            break
                    except:
                        continue

                if menu:
                    return menu

            except Exception as e:
                logger.debug(f"Menu interaction attempt failed: {e}")
                pass

            try:
                await self.page.mouse.wheel(0, 200)
            except:
                pass
            await self.context.wait_for_timeout(200)

        return None
    
    async def _select_export_type_excel_by_id(self):
        """
        Selects Export type = 'Microsoft Excel (.xlsx)' using the concrete IDs you provided:
          - Combobox:  #Dropdown486
          - Listbox:   #Dropdown486-list
        Tries both top-level page and iframe contexts (callouts may render in either).
        """
        # Ensure dialog is visible
        dlg = self.page.locator(".ms-Dialog-main").last
        await dlg.wait_for(state="visible", timeout=10000)

        # 1) Open the 'Export type' combobox by its concrete ID (inside the dialog)
        dd = dlg.locator("#Dropdown486")
        if await dd.count() == 0:
            # Fallback: find via label 'Export type' → nearest combobox
            lab = dlg.locator("label#Dropdown486-label, label:has-text('Export type')")
            if await lab.count() > 0:
                dd = lab.nth(0).locator("xpath=following::*[@role='combobox'][1]")
        if await dd.count() == 0:
            raise RuntimeError("Export type combobox (#Dropdown486) not found.")

        await dd.first.click()

        # 2) Wait for the dropdown list to appear (can render in page OR frame)
        candidates = [self.page, self.context if self.context != self.page else None]
        found_list = None
        for ctx in candidates:
            if not ctx:
                continue
            lst = ctx.locator("#Dropdown486-list")
            if await lst.count() > 0:
                try:
                    await lst.first.wait_for(state="visible", timeout=5000)
                    found_list = lst.first
                    break
                except:
                    pass
        if not found_list:
            # As a fallback, try role=listbox in either context
            for ctx in candidates:
                if not ctx:
                    continue
                role_list = ctx.get_by_role("listbox")
                if await role_list.count() > 0:
                    found_list = role_list.first
                    break
        if not found_list:
            raise RuntimeError("Export type list (#Dropdown486-list) did not appear.")

        # 3) Click the 'Microsoft Excel (.xlsx)' option by text
        # Prefer exact list scope first, then fall back to role=option search.
        opt = found_list.locator("button:has-text('Microsoft Excel (.xlsx)')").first
        if await opt.count() == 0:
            # Fallback: role-based in same context
            ctx = found_list.page if hasattr(found_list, "page") else self.page
            opt = ctx.get_by_role("option", name=re.compile(r"Microsoft\s+Excel\s*\(\.xlsx\)", re.I)).first

        if await opt.count() == 0:
            raise RuntimeError("Export type option 'Microsoft Excel (.xlsx)' not found.")

        await opt.click()

    async def _click_commandbar_export(self):
        """
        Click the CommandBar 'Export' button (primary or overflow).
        Tries top-level page first, then iframe context.
        """
        async def _try_in_ctx(ctx) -> bool:
            if not ctx:
                return False
            try:
                # Make sure we’re at the top so the bar is visible
                try:
                    await ctx.evaluate("() => window.scrollTo(0, 0)")
                except:
                    pass

                # Find a menubar
                bar = ctx.get_by_role("menubar").first
                if await bar.count() == 0:
                    # Fallback to known CommandBar container
                    bar = ctx.locator("div[title='CommandBar'], .ms-CommandBar").first
                    if await bar.count() == 0:
                        return False

                # 1) Try visible primary command: role=menuitem name=Export
                export_btn = bar.get_by_role("menuitem", name=re.compile(r"^export$", re.I)).first
                if await export_btn.count() > 0 and await export_btn.is_visible():
                    await export_btn.click()
                    return True

                # Fallback by title attribute
                export_btn2 = bar.locator("button[title='Export']").first
                if await export_btn2.count() > 0 and await export_btn2.is_visible():
                    await export_btn2.click()
                    return True

                # 2) Try overflow (“More …”) then select Export from the menu
                overflow_triggers = [
                    bar.get_by_role("button", name=re.compile(r"(more( commands| options)?|see more)", re.I)).first,
                    bar.locator("button[aria-haspopup='true']").filter(
                        has=bar.locator("i[data-icon-name='ChevronDown']")
                    ).first,
                ]
                for trigger in overflow_triggers:
                    if await trigger.count() > 0:
                        try:
                            await trigger.click()
                            # Wait for a menu/callout to appear in this context
                            menu = ctx.get_by_role("menu").last
                            if await menu.count() == 0:
                                menu = ctx.locator(".ms-ContextualMenu, .ms-Callout").last
                            await menu.wait_for(state="visible", timeout=5000)

                            # Look for Export inside the overflow
                            item = menu.get_by_role("menuitem", name=re.compile(r"^export$", re.I)).first
                            if await item.count() == 0:
                                # fallback by text
                                item = menu.get_by_text(re.compile(r"^export$", re.I)).first
                            if await item.count() > 0:
                                await item.click()
                                return True
                        except:
                            # try next trigger type if present
                            pass

                return False
            except:
                return False

        # Try top-level page, then iframe context
        if await _try_in_ctx(self.page):
            return
        if await _try_in_ctx(self.context if self.context != self.page else None):
            return

        raise RuntimeError("CommandBar 'Export' button not found/visible.")

    

    async def _select_combobox(self, label_text_regex: str, option_text_regex: str, timeout_ms: int = 10000):
        """
        Opens a Fluent UI combobox by its accessible label and selects the option that matches option_text_regex.
        Uses Playwright role selectors for stability across CSS class changes.
        """
        # Use top-level page for the modal/dialog content
        cb = self.page.get_by_role("combobox", name=re.compile(label_text_regex, re.I))
        await cb.first.click()
        opt = self.page.get_by_role("option", name=re.compile(option_text_regex, re.I)).first
        await opt.click()
        
    async def _select_all_fields_in_dialog(self):
        """
        Select ALL fields fast.
        - Opens 'Select fields' dropdown
        - In a few JS passes, clicks any unchecked options currently rendered
        - Scrolls to materialize more (virtualized lists), repeats up to ~8 passes
        - Closes the callout
        Target runtime: ~0.5–3s instead of ~30s.
        """
        dlg = self.page.locator(".ms-Dialog-main").last
        await dlg.wait_for(state="visible", timeout=10000)

        # 0) Small settle after enabling the checkbox (Fluent needs a tick to mount dropdown)
        await self.page.wait_for_timeout(80)

        # 1) Resolve the "Select fields" combobox via its label
        name_rx = re.compile(r"^select\s*fields$", re.I)
        label = dlg.locator("label.ms-Dropdown-label", has_text=name_rx).first
        await label.wait_for(state="visible", timeout=6000)

        label_id = await label.get_attribute("id")              # e.g., 'Dropdown389-label'
        base_id  = (label_id or "").replace("-label", "")       # -> 'Dropdown389'

        combo = dlg.locator(f"div#{base_id}[role='combobox']").first if base_id else None
        if not combo or await combo.count() == 0:
            combo = label.locator("xpath=following-sibling::*[@role='combobox'][1]").first

        if await combo.count() == 0:
            logger.warning("'Select fields' combobox not found.")
            return

        # 2) Open if not already expanded
        if (await combo.get_attribute("aria-expanded") or "false") != "true":
            await combo.click()
            await self.page.wait_for_timeout(40)

        # 3) Find the listbox (callout is at page root)
        list_id = (await combo.get_attribute("aria-controls")) or (await combo.get_attribute("aria-owns")) or (f"{base_id}-list" if base_id else None)
        listbox = self.page.locator(f"#{list_id}").first if list_id else None
        if not listbox or await listbox.count() == 0:
            listbox = self.page.get_by_role("listbox").last
        await listbox.wait_for(state="visible", timeout=4000)

        # 4) FAST: Toggle all visible unchecked items in a few JS passes, scrolling between passes
        # Works for the structure you shared: div.ms-Checkbox.ms-Dropdown-item ... <label> ... <input type="checkbox">
        passes = 0
        total_changed = 0
        max_passes = 8                   # bounds total time
        scroll_step_js = "el => el.scrollBy(0, el.clientHeight)"  # page-by-page

        # Ensure we start from the top of the list
        try:
            await listbox.evaluate("el => { el.scrollTop = 0; }")
        except Exception:
            pass

        while passes < max_passes:
            passes += 1

            # One JS pass: click labels for any unchecked items currently rendered
            changed = await listbox.evaluate("""
                (root) => {
                    let changed = 0;
                    const evOpts = { bubbles: true, cancelable: true };
                    const items = root.querySelectorAll('.ms-Checkbox.ms-Dropdown-item');
                    items.forEach(host => {
                        const input = host.querySelector('input[type="checkbox"]');
                        if (input && !input.checked) {
                            // Prefer clicking label so React/Fluent updates selection state
                            const lbl = host.querySelector('label');
                            if (lbl) lbl.click();
                            else host.click();
                            changed++;
                        }
                    });
                    return changed;
                }
            """) or 0
            total_changed += changed

            # If nothing changed this pass, check if any remain unchecked; if not, we can stop
            remaining = 0
            try:
                remaining = await listbox.locator('input[type="checkbox"]:not(:checked)').count()
            except Exception:
                remaining = 0

            if remaining == 0:
                break

            # Scroll to materialize more items (virtualization)
            try:
                await listbox.evaluate(scroll_step_js)
            except Exception:
                # Fallback to keyboard scrolling
                try:
                    await listbox.focus()
                    await self.page.keyboard.press("PageDown")
                except Exception:
                    pass

            # Tiny settle between passes
            await self.page.wait_for_timeout(40)

        logger.info(f"[SelectFields] Passes={passes}, toggled={total_changed}")

        # 5) Close callout so Export button is reachable
        try:
            await self.page.keyboard.press("Escape")
        except Exception:
            try:
                await combo.click()
            except Exception:
                pass

    
    async def _enable_large_exports_checkbox(self):
        """
        Enable 'Enable large exports' immediately using the fastest proven path:
        get_by_label(...).set_checked(True, force=True).
        No retries, no alternative strategies.
        """
        dlg = self.page.locator(".ms-Dialog-main").last
        await dlg.wait_for(state="visible", timeout=5000)

        name_rx = re.compile(r"enable\s+large\s+exports", re.I)

        # Locate by accessible label (dialog first, then page scope because Fluent callouts can escape)
        target = dlg.get_by_label(name_rx).first
        if await target.count() == 0:
            target = self.page.get_by_label(name_rx).first

        if await target.count() == 0:
            logger.warning("[EnableLargeExports] Checkbox not found by label")
            return

        # Fast path: if already checked, bail
        try:
            if await target.is_checked():
                logger.info("[EnableLargeExports] Already checked")
                return
        except Exception:
            pass

        # Short-circuit: the method you confirmed is reliable
        await target.set_checked(True, force=True)

        # Verify once and return
        try:
            if await target.is_checked():
                logger.info("[EnableLargeExports] Checked via label.set_checked(force=True)")
                return
        except Exception:
            pass

        logger.warning("[EnableLargeExports] set_checked(force=True) did not stick")
    
    async def _click_export_and_download(self, data_dir: str, file_base: str):
        """
        Click 'Export' and save the download as <file_base>.xlsx (overwrite if exists).
        """
        os.makedirs(data_dir, exist_ok=True)

        dlg = self.page.locator(".ms-Dialog-main").last
        await dlg.wait_for(state="visible", timeout=10000)

        export_btn = dlg.get_by_role("button", name=re.compile(r"^export$", re.I)).first
        await export_btn.wait_for(state="visible", timeout=10000)

        # Wait until enabled (validation)
        for _ in range(40):  # ~20s @ 500ms
            try:
                if await export_btn.is_enabled():
                    break
            except:
                pass
            await self.page.wait_for_timeout(500)
        else:
            raise RuntimeError("Export button did not become enabled in time.")

        # Download
        async with self.page.expect_download() as dl_info:
            await export_btn.click()
        dl: Download = await dl_info.value

        # Force deterministic filename: "<file_base>.xlsx"
        suggested = dl.suggested_filename or "export.xlsx"
        _, ext = os.path.splitext(suggested)
        ext = ext if ext.lower() in (".xlsx", ".csv") else ".xlsx"

        safe_base = _sanitize_filename(file_base) or "visual"
        out_path = os.path.join(data_dir, f"{safe_base}{ext}")

        # Overwrite if exists
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass

        await dl.save_as(out_path)
        logger.info(f"Export saved → {out_path}")

        # Dismiss toast so menubar is clickable again
        try:
            await self._dismiss_export_toast(timeout_ms=6000)
        except Exception:
            pass

        await self.page.wait_for_timeout(150)
        return out_path

    
    async def _dismiss_export_toast(self, timeout_ms: int = 6000):
        """
        Closes the Fluent UI success toast (ms-MessageBar) that appears after export.
        Uses close button when present; falls back to ESC or hard remove via JS.
        """
        # The toast renders at page root (outside iframe/dialog)
        toast_container = self.page.locator(".ms-MessageBar")

        # If no toast shows up quickly, skip
        try:
            await toast_container.first.wait_for(state="visible", timeout=timeout_ms)
        except Exception:
            return

        # There can be multiple stacked toasts; attempt to close all
        try:
            count = await toast_container.count()
        except Exception:
            count = 1

        for i in range(count):
            toast = toast_container.nth(i)
            try:
                # Prefer the explicit dismissal button
                close_btn = toast.locator(
                    "button.ms-MessageBar-dismissal, "
                    "button[title='Close'], "
                    "button[aria-label='Close']"
                ).first

                if await close_btn.count() > 0:
                    try:
                        await close_btn.click()
                    except Exception:
                        try:
                            await close_btn.click(force=True)
                        except Exception:
                            pass
                else:
                    # Fallback: try Escape
                    try:
                        await self.page.keyboard.press("Escape")
                    except Exception:
                        pass

                # Verify removal/hide; hard-remove as last resort
                try:
                    await toast.wait_for(state="detached", timeout=1500)
                except Exception:
                    try:
                        await self.page.evaluate(
                            "() => { document.querySelectorAll('.ms-MessageBar').forEach(n => n.remove()); }"
                        )
                    except Exception:
                        pass

            except Exception:
                # If anything odd, just try nuking all toasts
                try:
                    await self.page.evaluate(
                        "() => { document.querySelectorAll('.ms-MessageBar').forEach(n => n.remove()); }"
                    )
                except Exception:
                    pass

        # tiny settle so menubar becomes interactive again
        await self.page.wait_for_timeout(100)


    async def _export_via_menubar(self, safe_page_name: str, visual_title: str, data_dir: str, file_base: str, visual_index: int):
        """
        Menubar export flow; names file as <file_base>.xlsx (overwrites).
        If visual_title is empty, selects the visual by index in the 'Select visual' dropdown.
        """
        await self._click_commandbar_export()

        dlg = self.page.locator(".ms-Dialog-main").last
        await dlg.wait_for(state="visible", timeout=10000)

        # Export with → Current view
        await self._select_combobox(r"^export with", r"^current view$")

        # Export type → Microsoft Excel (.xlsx)
        await self._select_export_type_excel_by_id()

        # Select visual → by title if we have one; otherwise by index
        if (visual_title or "").strip():
            await self._select_combobox(r"^select visual", re.escape(visual_title))
        else:
            # Open combobox
            combo = self.page.get_by_role("combobox", name=re.compile(r"^select\s+visual", re.I)).first
            await combo.click()
            await self.page.wait_for_timeout(50)

            # The listbox may render at page root
            listbox = self.page.get_by_role("listbox").last
            await listbox.wait_for(state="visible", timeout=4000)

            options = listbox.get_by_role("option")
            count = await options.count()
            if count == 0:
                logger.warning("[Export] No options in 'Select visual' list.")
            else:
                idx = visual_index if visual_index < count else count - 1
                await options.nth(idx).click()

        # Enable large exports
        await self._enable_large_exports_checkbox()

        # Select fields → all
        await self._select_all_fields_in_dialog()

        # Export and save with deterministic name
        await self._click_export_and_download(data_dir, file_base)

        await self.page.wait_for_timeout(300)



    async def _export_visuals_on_current_page(self, safe_page_name: str, allowed_visuals: Set[str]):
        allowed_visuals = allowed_visuals or set()
        if not allowed_visuals:
            logger.info("No allowed visuals configured for this page; skipping.")
            return

        # Try different selectors for visual containers
        selectors = [
            ".visualContainer[role='group']",  # Standard Power BI
            ".visual-container",               # EmbedFast common
            "[class*='visual'][role='group']", # Generic visual container
            "[class*='visual-container']",     # Another common pattern
        ]

        containers = []
        for selector in selectors:
            try:
                await self.context.wait_for_selector(selector, timeout=5000)
                all_containers = await self.context.query_selector_all(selector)
                
                # Check visibility
                for c in all_containers:
                    try:
                        if await c.is_visible():
                            containers.append(c)
                    except:
                        continue
                
                if containers:
                    logger.info(f"Found visuals using selector: {selector}")
                    break
            except Exception as e:
                logger.debug(f"Selector '{selector}' failed: {e}")
                continue

        logger.info(f"Visual containers found: {len(all_containers)} | Visible containers: {len(containers)}")
        if not containers:
            logger.warning("No visible visuals found on this page.")
            return
        
        page_dir = os.path.join(self.download_dir, safe_page_name)
        visuals_dir = os.path.join(page_dir, "visuals")
        data_dir    = os.path.join(page_dir, "data")
        os.makedirs(visuals_dir, exist_ok=True)
        os.makedirs(data_dir, exist_ok=True)

        processed_any = False
        for i, container in enumerate(containers):
            try:
                human_title = await self._get_visual_title_for_matching(container)
                export_all = "__ALL__" in allowed_visuals or "all" in allowed_visuals
                if not export_all and _norm(human_title) not in allowed_visuals:
                    continue  # not in whitelist

                processed_any = True
                visual_name = await self._get_visual_name_for_files(container, i)

                # Crop target: inner 'visualWrapper' preferred
                wrapper = await container.query_selector("[data-testid='visual-style'].visualWrapper, .visualWrapper")
                target_for_shot = wrapper or container

                # ---- Screenshot (OVERWRITE) ----
                img_path = os.path.join(visuals_dir, f"{visual_name}.png")
                await target_for_shot.screenshot(path=img_path)
                logger.info(f"[{self.config_report_name} | {safe_page_name} | {human_title}] Screenshot → {img_path}")

                                # ---- Export via CommandBar (NEW FLOW) ----
                try:
                    await self._export_via_menubar(safe_page_name, human_title, data_dir, visual_name, i)
                except Exception as e:
                    logger.warning(f"[{self.config_report_name} | {safe_page_name} | {human_title}] Menubar export failed: {e}")

                # ---- Export data ----
                # menu = await self._open_menu_for_container(container, retries=3)
                # if not menu:
                #     logger.warning(f"[{human_title}] Could not open menu; skipping export.")
                #     continue

                # clicked = False
                # items = await menu.query_selector_all(".pbi-menu-item-text-container, [role='menuitem']")
                # for it in items:
                #     txt = (await it.inner_text() or "").strip().lower()
                #     if "export data" in txt:
                #         await it.click()
                #         clicked = True
                #         break

                # if not clicked:
                #     logger.warning(f"[{human_title}] 'Export data' not found; skipping export.")
                #     try:
                #         await self.page.keyboard.press("Escape")
                #     except:
                #         pass
                #     continue

                # try:
                #     await self.context.wait_for_selector(".pbi-modern-button", timeout=6000)
                #     buttons = await self.context.query_selector_all(".pbi-modern-button")
                #     did_export = False
                #     for b in buttons:
                #         bt = (await b.inner_text() or "").lower()
                #         if "export" in bt:
                #             async with self.page.expect_download() as dl_info:
                #                 await b.click()
                #             dl: Download = await dl_info.value
                #             filename = dl.suggested_filename
                #             out_path = os.path.join(data_dir, filename)
                #             # ---- OVERWRITE behavior ----
                #             await dl.save_as(out_path)
                #             logger.info(f"[{self.config_report_name} | {safe_page_name} | {human_title}] Export → {out_path}")
                #             did_export = True
                #             break

                #     if not did_export:
                #         logger.warning(f"[{human_title}] Export dialog buttons not found/clicked.")
                # except Exception as e:
                #     logger.warning(f"[{human_title}] Export failed: {e}")
                #     try:
                #         await self.page.keyboard.press("Escape")
                #     except:
                #         pass
                #     continue

                await self.page.wait_for_timeout(150)

            except Exception as e:
                logger.error(f"Failed to process a visual: {e}")
                continue

        if not processed_any:
            logger.info("No visuals matched the whitelist on this page.")

# -------------------- ENTRY -------------------- #
if __name__ == "__main__":
    runner = ReportsRunner()
    asyncio.run(runner.run())